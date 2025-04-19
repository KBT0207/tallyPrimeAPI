import pandas as pd
import numpy as np
from logging_config import logger
from io import BytesIO
from xlwings import view
import os
import io
import time

from utils.common_utils import kaybee_exports_currency, receivables_comp_codes
import requests
# from database.sql_connector import kbe_connector, kbe_connection, kbe_engine
from datetime import datetime
import openpyxl


def get_filename_tally(path:str):
    return path.split("\\")[-1].rsplit("_", 2)  [-2]

def get_compname_tally(path:str):
    return path.split("\\")[-1].rsplit("_", 2)[0]


def get_date_tally(path:str):
    return path.split("\\")[-1].split("_")[-1].removesuffix(".xlsx")


pd.set_option('future.no_silent_downcasting', True)

def apply_transformation(file_path, material_centre_name: str) -> pd.DataFrame:
    if os.path.basename(file_path).startswith('~$'):
        logger.info(f"Skipping temporary file: {file_path}")
        return None
    fcy_mc_list = ["FCY_Frexotic", "FCY_KBE", "FCY_KBEIPL", "FCY_KBAIPL", 'FCY_Orbit','FCY_Freshnova']
    if material_centre_name in fcy_mc_list:
        return fcy_helper_apply_transformation(file_path, material_centre_name)
    else:
        return helper_apply_transformation(file_path, material_centre_name)

def apply_register_transformation(file_path, material_centre_name: str) -> pd.DataFrame:
    if os.path.basename(file_path).startswith('~$'):
        logger.info(f"Skipping temporary file: {file_path}")
        return None
    fcy_mc_list = ["FCY_Frexotic", "FCY_KBE", "FCY_KBEIPL", "FCY_KBAIPL", 'FCY_Orbit','FCY_Freshnova']
    if material_centre_name in fcy_mc_list:
        return fcy_helper_apply_register_transformation(file_path, material_centre_name)
    else:
        return helper_apply_register_transformation(file_path, material_centre_name)

def helper_apply_transformation(file_path, material_centre_name:str) -> pd.DataFrame:
    
    mc = material_centre_name.replace('_', " ")
    fcy_mc_list = ["FCY Frexotic", "FCY KBE", "FCY KBEIPL", "FCY KBAIPL", 'FCY Freshnova'] 
    try:
        df = pd.read_excel(file_path, skipfooter= 1, header=None) 
        date_row = df[df.iloc[:, 0] == 'Date'].index[0]
        df = df.iloc[date_row:].reset_index(drop=True)
        df.columns = df.iloc[0]
        df = df.iloc[1:]
        df = df.drop(index=1)
    except FileNotFoundError as e:
        logger.info(e)
        logger.warning(f"Excel File not found in the given {file_path}: {e}")
    if df.empty:
        logger.warning(f"Empty Excel File of {get_compname_tally(file_path)} and report {get_filename_tally(file_path)}")
        return None
    df.columns = df.columns.str.lower().str.replace(" ", "_").str.replace(".", "")
    df = df.rename(columns= {"vch_type": "voucher_type", "vch_no": "voucher_no"})
    df['voucher_no'] = np.where((df['date'].notnull()) & (df['voucher_no'].isnull()), 'Blank', df['voucher_no'])

    material_center = mc

    currency_name = kaybee_exports_currency.get(material_center,None)

    df.loc[:, ["credit", "debit"]] = df.loc[:, ["credit", "debit"]].fillna(0)

    df["material_centre"] = material_center
    df["currency"] = currency_name
    
    df["fcy"] = df["material_centre"].apply(lambda x: "Yes" if x in fcy_mc_list else "No")
    
    df["particulars"] = df["particulars"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)
    df["voucher_no"] = df["voucher_no"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)

    df["voucher_no"] = df["voucher_no"].fillna(df["particulars"])
    df = df.loc[~df["particulars"].isna()]

    return df

def fcy_helper_apply_transformation(file_path, material_centre_name):
    try:
        fcy_mc_list = ["FCY Frexotic", "FCY KBE", "FCY KBEIPL", "FCY KBAIPL", 'FCY Orbit','FCY Freshnova']
        mc = material_centre_name.replace('_', " ")

        # Load the workbook
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
        except Exception as e:
            logger.info(f"Error loading the Excel file: {e}")
            return None

        # Currency symbols and their corresponding codes
        currency_map = {
            "AU$": "AUD",
            "$": "USD",
            "€": "EUR",
            "£": "GBP",
            "₹": "INR",
            "AUD": "AUD",
            "CAD": "CAD",
        }

        # Initialize variables for columns
        header_row = None
        debit_column_index = None
        credit_column_index = None

        # Iterate through the rows to find the relevant columns
        try:
            for row_id, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if 'Date' in row:
                    header_row = row_id
                if 'Debit' in row:
                    debit_column_index = row.index('Debit') + 1
                if 'Credit' in row:
                    credit_column_index = row.index('Credit') + 1

                # Exit once all required columns are found
                if header_row and debit_column_index and credit_column_index:
                    break
        except Exception as e:
            logger.info(f"Error processing rows for columns: {e}")
            return None

        # Function to check the format and extract the currency symbol
        def extract_currency_from_format(cell):
            """Extracts the currency symbol from the cell's number format."""
            try:
                number_format = cell.number_format
                # Checking for AU$ in number format (if it's used in formatting)
                if 'AU$' in number_format or 'AUD' in number_format or 'A$' in number_format:
                    return 'AU$'
                if '$' in number_format:
                    return '$'
                elif '€' in number_format:
                    return '€'
                elif '£' in number_format:
                    return '£'
                elif '¥' in number_format:
                    return '¥'
                elif 'CAD' in number_format:
                    return 'CAD'
                return None
            except Exception as e:
                logger.info(f"Error extracting currency from format: {e}")
                return None

        # Function to extract currency from a string value
        def extract_currency_from_string(value):
            """Extracts the currency symbol from a string value."""
            try:
                value_str = str(value).strip()
                for symbol, code in currency_map.items():
                    if symbol in value_str:
                        return symbol  # Return the exact symbol found
                return None
            except Exception as e:
                logger.info(f"Error extracting currency from string: {e}")
                return None

        # Add a new column for the unified currency (Debit or Credit) in memory
        currency_data = []

        # Now iterate through the rows starting from the row after the header
        try:
            for row in range(header_row + 1, ws.max_row + 1):
                debit_cell = ws.cell(row=row, column=debit_column_index)
                credit_cell = ws.cell(row=row, column=credit_column_index)

                # First check if the currency symbol is extracted from number format in Debit column
                debit_currency_symbol = extract_currency_from_format(debit_cell)
                credit_currency_symbol = extract_currency_from_format(credit_cell)

                # If no symbol is found in the number format, check the string value for Debit column
                if not debit_currency_symbol:
                    debit_currency_symbol = extract_currency_from_string(debit_cell.value)
                if not credit_currency_symbol:
                    credit_currency_symbol = extract_currency_from_string(credit_cell.value)

                # If currency is found in Debit, use it; otherwise, use Credit
                if debit_currency_symbol:
                    currency_code = currency_map.get(debit_currency_symbol, None)
                elif credit_currency_symbol:
                    currency_code = currency_map.get(credit_currency_symbol, None)
                else:
                    currency_code = None

                # Collect the row data and add the currency column
                row_data = [cell.value for cell in ws[row]]
                row_data.append(currency_code)  # Add the currency column
                currency_data.append(row_data)
        except Exception as e:
            logger.info(f"Error processing rows for currency extraction: {e}")
            return None

        # Create a Pandas DataFrame with the extracted data
        try:
            df = pd.DataFrame(currency_data, columns=[cell.value for cell in ws[header_row]] + ["Currency"])
            df.columns = df.columns.str.lower().str.replace(" ", "_").str.replace(".", "")
        except Exception as e:
            logger.info(f"Error creating DataFrame: {e}")
            return None

        # Data cleaning and transformation
        try:
            df['currency'] = df['currency'].fillna('Unknown')
            df['date'] = pd.to_datetime(df['date'], errors='coerce').dt.strftime('%Y-%m-%d')
            df = df.drop(index=[0, df.index[-1]])
            df = df.reset_index(drop=True)
            df[['debit', 'credit']] = df[['debit', 'credit']].fillna(0)
            df = df.rename(columns={"vch_type": "voucher_type", "vch_no": "voucher_no"})
            df['voucher_no'] = np.where((df['date'].notnull()) & (df['voucher_no'].isnull()), 'Blank', df['voucher_no'])
            df['material_centre'] = mc
            df['fcy'] = df['material_centre'].apply(lambda x: 'Yes' if x in fcy_mc_list else 'No')

            df['particulars'] = df['particulars'].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)
            df['voucher_no'] = df['voucher_no'].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)
            df['voucher_no'] = df['voucher_no'].fillna(df['particulars'])

            df = df.loc[~df['particulars'].isna()]

            df = df[['date', 'particulars', "voucher_type", 'voucher_no', 'debit', 'credit', 'material_centre', "currency", 'fcy']]

            df['debit'] = pd.to_numeric(df['debit'].astype(str).str.replace(',', ''), errors='coerce')
            df['credit'] = pd.to_numeric(df['credit'].astype(str).str.replace(',', ''), errors='coerce')

        except Exception as e:
            logger.info(f"Error in data cleaning and transformation: {e}")
            return None
        
        return df

    except Exception as e:
        logger.info(f"An error occurred: {e}")
        return None

def helper_apply_register_transformation(file_path, material_centre_name) -> pd.DataFrame:
    mc = material_centre_name.replace('_', " ")
    fcy_columns = ["FCY Frexotic", "FCY KBE", "FCY KBEIPL", "FCY KBAIPL", 'FCY Orbit','FCY Freshnova']
    try:
        df = pd.read_excel(file_path, skipfooter= 1, header=None)
        date_row = df[df.iloc[:, 0] == 'Date'].index[0]
        df = df.iloc[date_row:].reset_index(drop=True)
        df.columns = df.iloc[0]
        df = df.iloc[1:]
        df = df.drop(index=1)
    except FileNotFoundError as e:
        logger.info(e)
        logger.warning(f"Excel File not found in the given {file_path}: {e}")
    if df.empty:
        logger.warning(f"Empty Excel File of {get_compname_tally(file_path)} and report {get_filename_tally(file_path)}")
        return None
    df.columns = df.columns.str.lower().str.replace(" ", "_").str.replace(".", "")
    
    currency_name = kaybee_exports_currency.get(mc)
    df = df.rename(columns= {"vch_no": "voucher_no"})
    
    material_center = mc
    df["material_centre"] = material_center
    df["currency"] = currency_name
    df["fcy"] = "Yes" if material_center in fcy_columns else "No"
    
    df["particulars"] = df["particulars"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)
    df["voucher_no"] = df["voucher_no"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)
    df['voucher_no'] = np.where((df['date'].notnull()) & (df['voucher_no'].isnull()), 'Blank', df['voucher_no'])
    df.loc[:,["date",'voucher_no']] = df.loc[:,["date",'voucher_no']].ffill()
    df['date'] = pd.to_datetime(df["date"])
    df.loc[:,['credit', 'debit']] = df.loc[:,['credit', 'debit']].fillna(0)
    
    df['amount'] = np.where((df['credit'] != 0), df['credit'], df['debit'])
    df['amount'] = np.where((df['debit'] != 0), df['debit'], df['amount'])

    df["amount_type"] = np.where(df['credit'] != 0, 'credit', 'debit')
    df = df.drop(columns= ["vch_type", "debit", "credit"])
    df['particulars'] = np.where(df['particulars'] == "(cancelled)", "Cancelled", df['particulars'])
    df = df[['date', 'particulars', 'voucher_no','material_centre', 'amount', 'amount_type',"currency",'fcy',]]
    df = df.loc[~df["particulars"].isna()]
    return df

def fcy_helper_apply_register_transformation(file_path, material_centre_name):
    mc = material_centre_name.replace('_', " ")
    fcy_columns = ["FCY Frexotic", "FCY KBE", "FCY KBEIPL", "FCY KBAIPL", 'FCY Orbit','FCY Freshnova']

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except Exception as e:
        logger.info(f"Error loading the Excel file: {e}")
        return None

    currency_map = {"AU$": "AUD", "$": "USD", "€": "EUR", "£": "GBP", "₹": "INR", "AUD": "AUD", "CAD": "CAD", "A$": "AUD"}

    header_row = None
    credit_column_index = None
    debit_column_index = None
    particular_column_index = None

    def extract_currency_from_format(cell):
        """Extracts the currency symbol from the cell's number format."""
        try:
            number_format = cell.number_format
            if 'AU$' in number_format or 'AUD' in number_format or 'A$' in number_format:
                return 'AU$'
            if '$' in number_format:
                return '$'
            elif '€' in number_format:
                return '€'
            elif '£' in number_format:
                return '£'
            elif '¥' in number_format:
                return '¥'
            elif 'CAD' in number_format:
                return 'CAD'
            return None
        except Exception as e:
            logger.info(f"Error extracting currency from format: {e}")
            return None

    def extract_currency_from_string(value):
        """Extracts the currency symbol from a string value."""
        try:
            value_str = str(value).strip()
            for symbol, code in currency_map.items():
                if symbol in value_str:
                    return symbol  # Return the exact symbol found
            return None
        except Exception as e:
            logger.info(f"Error extracting currency from string: {e}")
            return None

    try:
        for row_id, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if 'Date' in row:
                header_row = row_id
            if 'Credit' in row:
                credit_column_index = row.index('Credit') + 1
            if 'Debit' in row:
                debit_column_index = row.index('Debit') + 1
            if 'Particulars' in row:
                particular_column_index = row.index('Particulars') + 1
            if header_row and particular_column_index and credit_column_index and debit_column_index:
                break

        ws.delete_cols(particular_column_index + 1)
        ws.cell(header_row, particular_column_index + 1, "rate")
        ws.cell(header_row, credit_column_index, "currency")
        ws.cell(header_row, credit_column_index + 1, "rate_currency")

    except Exception as e:
        logger.info(f"Error processing rows for columns: {e}")
        return None

    try:
        for row in range(header_row + 1, ws.max_row + 1):
            debit_cell = ws.cell(row=row, column=debit_column_index - 1)
            credit_cell = ws.cell(row=row, column=credit_column_index - 1)
            rate_cell = ws.cell(row=row, column=particular_column_index + 1)
            rate_currency_symbol = extract_currency_from_format(rate_cell)
            if not rate_currency_symbol:
                rate_currency_symbol = extract_currency_from_string(rate_cell)

            currency_code = extract_currency_from_format(credit_cell)
            if not currency_code:
                currency_code = extract_currency_from_string(credit_cell.value)
            if not currency_code:
                currency_code = extract_currency_from_format(debit_cell)
                if not currency_code:
                    currency_code = extract_currency_from_string(debit_cell.value)
                if not currency_code:
                    currency_code = None
            ws.cell(row=row, column=credit_column_index, value=currency_map.get(currency_code))
            ws.cell(row=row, column=credit_column_index + 1, value=currency_map.get(rate_currency_symbol))

    except Exception as e:
        logger.info(f"Error processing rows for currency extraction: {e}")
        return None

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    try:
        df = pd.read_excel(buffer, engine='openpyxl',skipfooter=1)
        date_row = df[df.iloc[:, 0] == 'Date'].index[0]
        df = df.iloc[date_row:].reset_index(drop=True)
        df.columns = df.iloc[0]
        df = df.iloc[1:]
        df = df.drop(index=1)

        if df.empty:
            logger.info("The DataFrame is empty after processing.")
            return None
        logger.info("Data read successfully with pandas:")

        try:
            df.columns = df.columns.str.lower().str.replace(".", "").str.replace(" ", "_")

            df = df.rename(columns= {"vch_no": "voucher_no"})

            df['date'] = pd.to_datetime(df['date'], errors='coerce').dt.strftime('%Y-%m-%d')

            df['voucher_no'] = np.where((df['date'].notnull()) & (df['voucher_no'].isnull()), 'Blank', df['voucher_no'])

            cols_bfill = ['rate','rate_currency']
            df[cols_bfill] = df[cols_bfill].bfill(limit=1)

            df['date'] = df['date'].ffill()
            df['voucher_no'] = df['voucher_no'].ffill()
            df = df.dropna(subset=['debit', 'credit'], how='all')

            df.loc[:,['credit', 'debit']] = df.loc[:,['credit', 'debit']].fillna(0)

            df["particulars"] = df["particulars"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)
            df["voucher_no"] = df["voucher_no"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)
            df['amount'] = np.where((df['credit'] != 0), df['credit'], df['debit'])
            df['amount'] = np.where((df['debit'] != 0), df['debit'], df['amount'])
            df["amount_type"] = np.where(df['credit'] != 0, 'credit', 'debit')
            df.loc[df['currency'].isnull() & (df['amount_type'] == 'debit'), 'amount'] = df['amount'] / df['rate']
            df.loc[df['currency'].isnull() & (df['amount_type'] == 'credit'), 'amount'] = df['amount'] / df['rate']
            df['currency'] = df['currency'].fillna(df['rate_currency'])
            df = df.drop(columns= ["vch_type", "debit", "credit"])
            df['particulars'] = np.where(df['particulars'] == "(cancelled)", "Cancelled", df['particulars'])
            df["material_centre"] = mc
            df["fcy"] = df["material_centre"].apply(lambda x: "Yes" if x in fcy_columns else "No")
            df = df.loc[~df["particulars"].isna()]
            df["voucher_no"] = df["voucher_no"].fillna('Blank')
            df = df[['date', 'particulars' ,'voucher_no','material_centre', 'amount', 'amount_type',"currency",'fcy']]
            df['amount'] = pd.to_numeric(df['amount']).round(2)

            return df
        except Exception as e:
            logger.info(f"Error during DataFrame operations: {e}")
            return None

    except Exception as e:
        logger.info(f"Error reading the modified file with pandas: {e}")
        return None

def apply_accounts_transformation(file_path, material_centre_name) -> pd.DataFrame:
    mc = material_centre_name.replace('_', " ")
    currency_name = kaybee_exports_currency.get(mc)
    try:
        df = pd.read_excel(file_path, header=None)
        header_row = df[df.iloc[:, 0] == 'Sl. No.'].index[0]
        df = df.iloc[header_row:].reset_index(drop=True)
        df.columns = df.iloc[0]
        df = df.iloc[1:]
    except FileNotFoundError as e:
        logger.warning(f"Excel File not found in the given {file_path}: {e}")
    if df.empty:
        logger.warning(f"Empty Excel File of {get_compname_tally(file_path)} and report {get_filename_tally(file_path)}")
        return None
    df.columns = df.columns.str.lower().str.replace(" ", "_").str.replace(".", "")
    df = df.drop(columns="sl_no", axis='columns')
    
    df = df.rename(columns= {"name_of_ledger": "ledger_name", "state_name": "state", 
                            "gstin/un": "gst_no",
                            })

    df["material_centre"] = mc
    df["currency"] = currency_name

    df["alias_code"] = df["ledger_name"].str.extract(pat= r'\(([^()]*?/[^()]*?)\)')
    df["alias_code"] = df["alias_code"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)

    df["ledger_name"] = df["ledger_name"].str.replace(r'\([^()]*?/[^()]*?\)$', '', regex= True).str.rstrip()
    df["ledger_name"] = df["ledger_name"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)

    df["under"] = df["under"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)
    df["gst_no"] = df["gst_no"].str.rstrip()
    df["opening_balance"] = df["opening_balance"].fillna(0)
    
    return df

def apply_kbe_accounts_transformation(file_path, material_centre_name) -> pd.DataFrame:
    mc = material_centre_name.replace('_', " ")
    try:
        df = pd.read_excel(file_path, header=None)
        header_row = df[df.iloc[:, 0] == 'Sl. No.'].index[0]
        df = df.iloc[header_row:].reset_index(drop=True)
        df.columns = df.iloc[0]
        df = df.iloc[1:]
    except FileNotFoundError as e:
        logger.warning(f"Excel File not found in the given {file_path}: {e}")
    if df.empty:
        logger.warning(f"Empty Excel File of {get_compname_tally(file_path)} and report {get_filename_tally(file_path)}")
        return None
    df.columns = df.columns.str.lower().str.replace(" ", "_").str.replace(".", "")
    df = df.drop(columns=["sl_no", "state_name"], axis='columns', errors= 'ignore')
    
    df = df.rename(columns= {"name_of_ledger": "ledger_name" })

    df["material_centre"] = mc

    df["alias_code"] = df["ledger_name"].str.extract(pat= r'\(([^()]*?/[^()]*?)\)')
    df["alias_code"] = df["alias_code"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)

    df["ledger_name"] = df["ledger_name"].str.replace(r'\([^()]*?/[^()]*?\)$', '', regex= True).str.rstrip()
    df["ledger_name"] = df["ledger_name"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)

    df["under"] = df["under"].str.replace('\n', '', regex=True).str.replace('_x0004_', '', regex=True)
    df["under"] = df["under"].str.lstrip()

    df["opening_balance"] = df["opening_balance"].fillna(0)

    df[['salesman', 'country']] = None
    df["credit_days"] = 0    
    return df

def apply_items_transformation(file_path:str) -> pd.DataFrame:
    try:
        df = pd.read_excel(file_path, header=None)
        header_row = df[df.iloc[:, 0] == 'Sl. No.'].index[0]
        df = df.iloc[header_row:].reset_index(drop=True)
        df.columns = df.iloc[0]
        df = df.iloc[1:]
    except FileNotFoundError as e:
        logger.info(e)
        logger.warning(f"Excel File not found in the given {file_path}: {e}")
    if df.empty:
        logger.warning(f"Empty Excel File of {get_compname_tally(file_path)} and report {get_filename_tally(file_path)}")
        return None
    df.columns = df.columns.str.lower().str.replace(" ", "_").str.replace(".", "")

    df = df.drop(columns="sl_no", axis='columns')
    
    df = df.rename(columns= {"name_of_item": "item_name",
                            })

    df["item_name"] = df["item_name"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)
    df["under"] = df["under"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)
    df["under"] = df["under"].replace('_x0004_','', regex=True)
    df.loc[:,['opening_qty', 'rate', 'opening_balance']] = df.loc[:,['opening_qty', 'rate', 'opening_balance']].fillna(0)
    
    return df

def apply_outstanding_balance_transformation(file_path, material_centre_name) -> pd.DataFrame:
    mc = 'test'
    try:
        df = pd.read_excel(file_path, header=None, skipfooter=1)
        credit_index = df.loc[df.eq('Credit').any(axis=1)].index[0]
        df = df.drop(range(credit_index)).reset_index(drop=True)
        df.columns = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)

    except FileNotFoundError as e:
        logger.warning(f"Excel File not found in the given {file_path}: {e}")
    if df.empty:
        logger.warning(f"Empty Excel File of {get_compname_tally(file_path)} and report {get_filename_tally(file_path)}")
        return None
    df.columns = df.columns.str.lower()

    df = df.rename(columns={np.nan: "particulars"})
    df["date"] = get_date_tally(path=file_path)
    df['date'] = pd.to_datetime(df['date'].str.removesuffix('.xlsx'), dayfirst=True)
    df["material_centre"] = mc

    df["particulars"] = df["particulars"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)
    df.loc[:, ["credit", "debit"]] = df[["credit", "debit"]].fillna(0)

    return df

def get_exchange_rate_in_inr(currency_code, date=None) -> dict | None:
    if date:
        api_url = f"https://api.frankfurter.app/{date}?base={currency_code}&symbols=INR"
    else:
        api_url = f"https://api.frankfurter.app/latest?base={currency_code}&symbols=INR"

    try:
        response = requests.get(api_url)
        response.raise_for_status()
        rates = response.json()
        
        # Extract rate and date
        if currency_code == 'INR':
            rate = 1
            rate_date = date if date else datetime.today().strftime('%Y-%m-%d')
        else:
            rate = rates.get('rates', {}).get('INR')
            rate_date = rates.get('date')

        # Return rate and date only if rate is successfully retrieved
        if rate is not None:
            return {"rate": rate, "date": rate_date}
        
    except requests.RequestException as e:
        # Log the exception if needed and return None
        logger.error(f"Failed to fetch exchange rate for {currency_code} on {date}: {e}")
        return None

def get_currency_code(format_string: str) -> str:
    """
    Extract standardized currency code from Excel number format string.
    
    Args:
        format_string (str): Excel number format string (e.g. '"0.00"CAD"', '"0.00"$"', '"0.00"€"')
    
    Returns:
        str: Standardized currency code (e.g. 'CAD', 'USD', 'EUR', 'INR')
    """
    # Handle empty or invalid input
    if not format_string or not isinstance(format_string, str):
        return 'INR' 

    # Remove the number format part and quotes
    cleaned_format = format_string.replace('"0.00"', '').replace('"', '')
    
    # Match symbols to currency codes
    if cleaned_format == 'CAD':
        return 'CAD'
    elif cleaned_format == '$':
        return 'USD'
    elif cleaned_format == '€':
        return 'EUR'
    elif cleaned_format == '₹':
        return 'INR'
    elif cleaned_format == '£':
        return 'GBP'
    elif cleaned_format == '¥':
        return 'JPY'
            
    # Default to INR if no match found
    return 'INR'


def apply_transformation_detailed(file_path: str, material_centre_name: str):
    mc = material_centre_name.replace('_', " ")
    report_name = get_filename_tally(file_path)
    fcy_list = [
        "FCY Frexotic",
        "FCY KBE",
        "FCY KBEIPL",
        "FCY Orbit",
        "FCY KBAIPL",
        'FCY Freshnova'
    ]

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except Exception as e:
        return None

    currency_map = {"AU$": "AUD", "$": "USD", "€": "EUR", "£": "GBP", "AUD": "AUD", "CAD": "CAD", "A$": "AUD"}

    def extract_currency_from_format(cell):
        try:
            number_format = cell.number_format
            if 'AU$' in number_format or 'AUD' in number_format or 'A$' in number_format:
                return 'AU$'
            if '$' in number_format:
                return '$'
            elif '€' in number_format:
                return '€'
            elif '£' in number_format:
                return '£'
            elif '¥' in number_format:
                return '¥'
            elif 'CAD' in number_format:
                return 'CAD'
            return None
        except Exception as e:
            return None

    def extract_currency_from_string(value):
        try:
            value_str = str(value).strip()
            for symbol, code in currency_map.items():
                if symbol in value_str:
                    return symbol
            return None
        except Exception as e:
            return None

    header_row = None
    rate_column_index = None

    try:
        if report_name == 'sales-detailed':

            try:
                for row_id, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if 'Date' in row:
                        header_row = row_id
                    if 'Rate' in row:
                        rate_column_index = row.index('Rate') + 1
                    if header_row and rate_column_index:
                        break
                if not header_row or not rate_column_index:
                    return None
                rate_currency_col = ws.max_column + 1
                ws.cell(row=header_row, column=rate_currency_col, value='rate_currency')
            except Exception as e:
                return None
            
        if report_name =='sales-return-detailed':
            try:
                for row_id, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    if 'Date' in row:
                        header_row = row_id
                    if 'Rate' in row:
                        rate_column_index = row.index('Rate') + 1
                    if header_row and rate_column_index:
                        break
                if not header_row or not rate_column_index:
                    return None
                rate_currency_col = ws.max_column + 1
                ws.cell(row=header_row, column=rate_currency_col, value='rate_currency')
            except Exception as e:
                return None
            
    except Exception as e:
        print(f"Report Type Error {e}")
        
            
    is_fcy = any(fcy in mc for fcy in fcy_list)
    if is_fcy:
        try:
            for row in range(header_row + 1, ws.max_row + 1):
                rate_cell = ws.cell(row=row, column=rate_column_index)
                rate_currency_symbol = extract_currency_from_format(rate_cell)
                if not rate_currency_symbol:
                    rate_currency_symbol = extract_currency_from_string(rate_cell.value)
                ws.cell(row=row, column=ws.max_column, value=currency_map.get(rate_currency_symbol))
            wb.close()
        except Exception as e:
            pass
    try:
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        fixed_cols_read = ['Date', 'Particulars', 'Voucher Type', 'Voucher No.', 'Voucher Ref. No.',
                        'Quantity', 'Rate', 'Value', 'Gross Total', 'rate_currency']
        df = pd.read_excel(buffer, engine="openpyxl", skipfooter=1, skiprows=header_row - 1,
                        usecols=fixed_cols_read)
        df.columns = df.columns.str.lower().str.replace(" ", "_", regex=True).str.replace(r"\.", "", regex=True)
        if 'item' not in df.columns:
            df['item'] = ''
        df['item'] = np.where(df['date'].isnull() & (df['item'] == ''), df['particulars'], df['item'])
        df['particulars'] = df['particulars'].mask(df['date'].isnull()).ffill()
        df['item'] = np.where(df['quantity'].isnull() & df['rate'].isnull() & df['value'].isnull(), 'No item',
                                df['item'])
        df['particulars'] = df['particulars'].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)
        df['voucher_no'] = df['voucher_no'].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)

        df['voucher_no'] = np.where(df['date'].notna() & df['voucher_no'].isnull(), 'Blank', df['voucher_no'])
        df['voucher_type'] = np.where(df['date'].notna() & df['voucher_type'].isnull(), 'Blank', df['voucher_type'])
        df['voucher_ref_no'] = np.where(df['date'].notna() & df['voucher_ref_no'].isnull(), 'Blank',
                                        df['voucher_ref_no'])

        df['date'] = pd.to_datetime(df['date'], errors='coerce').dt.strftime('%Y-%m-%d')

        df['other_deduction'] = np.where(df['value'].notna() & df['gross_total'].notna(),
                                        df['value'] - df['gross_total'], np.nan).astype(float)
        df['other_deduction'] = df['other_deduction'].ffill(limit=1)
        df['other_deduction'] = df['other_deduction'].fillna(0)
        cols_fill_zero = ['quantity', 'rate', 'value']
        df[cols_fill_zero] = np.where(df[cols_fill_zero].isnull(), 0, df[cols_fill_zero])

        cols_ffill = ['voucher_ref_no', 'voucher_no', 'voucher_type', 'date']
        df[cols_ffill] = df[cols_ffill].ffill()

        df['gross_total'] = df['gross_total'].ffill(limit=1)
        df['gross_total'] = df['gross_total'].fillna(0)
        df['value'] = np.where(
            (df['gross_total'] > 0) & (df['item'] == 'No item') & (df['value'] == 0),
            df['gross_total'],
            df['value'])

        df = df[df['item'].notna() & (df['item'].str.strip() != "")]

        currency_value = kaybee_exports_currency.get(mc)

        if currency_value and df['rate_currency'].isnull().any():
            df['rate_currency'] = df['rate_currency'].fillna(currency_value)

        df['rate_currency'] = df['rate_currency'].fillna("Unknown")

        df['material_centre'] = mc
        df['fcy'] = df['material_centre'].apply(lambda x: 'Yes' if any(fcy in x for fcy in fcy_list) else 'No')
        df = df.rename(columns={'quantity': 'qty'})
        if report_name =='sales-detailed':
            final_columns = [
                "date", "particulars", "voucher_type", "voucher_no", "voucher_ref_no",
                "item", "qty", "rate", "rate_currency", "value", "gross_total",
                "other_deduction", "material_centre", "fcy"
                ]
        if report_name == 'sales-return-detailed':
            final_columns = [
                "date", "particulars", "voucher_type", "voucher_no", "voucher_ref_no",
                "item", "qty", "rate", "rate_currency", "value", "gross_total", "material_centre", "fcy"
                ]
        df = df[final_columns]

        end = time.time()
        return df
        
    except Exception as e:
        return None



# def apply_kbe_outstanding_transformation(file_path, material_centre_name) -> pd.DataFrame:
#     try:
#         # Load the Excel file using pandas for data, but openpyxl for formats
#         df = pd.read_excel(file_path, skipfooter=1, header=None)
#         workbook = openpyxl.load_workbook(file_path, data_only=True)
#         sheet = workbook.active

#         # Identify the "Date" row to start processing data from
#         date_row = df[df.iloc[:, 0] == 'Date'].index[0]
        
#         # Store the original Excel row numbers before any filtering
#         df_original = df.iloc[date_row:].reset_index(drop=True)
#         df_original.columns = df_original.iloc[0]
#         df_original = df_original.iloc[1:]
#         df_original = df_original.drop(index=1)
        
#         # Add Excel row numbers before any filtering
#         df_original['excel_row'] = range(date_row + 3, date_row + 3 + len(df_original))

#     except FileNotFoundError as e:
#         logger.warning(f"Excel File not found in the given {file_path}: {e}")
#         return None

#     if df_original.empty:
#         logger.warning(f"Empty Excel File of {get_compname(file_path)} and report {get_filename(file_path)}")
#         return None

#     # Standardize column names
#     df_original.columns = df_original.columns.str.lower()

#     mc = kbe_outstanding_comp_codes[int(material_centre_name)]

#     # Rename columns as needed
#     df_original = df_original.rename(columns={"ref. no.": "voucher_no", "party's name": "particulars", 
#                                               "pending": "amount", "due on": "due_on", 
#                                               "overdue": "overdue_in_days" })

#     # Get raw cell formats before filtering
#     amount_col_idx = df_original.columns.get_loc("amount") + 1  # Adjust for openpyxl's 1-based indexing
#     formats_dict = {}
    
#     # Store formats with their Excel row numbers
#     for row_idx in df_original['excel_row']:
#         cell = sheet.cell(row=row_idx, column=amount_col_idx)
#         formats_dict[row_idx] = cell.number_format

#     # Filter rows where "amount" is greater than 0
#     df = df_original.loc[df_original["amount"] > 0].copy()
    
#     df["material_centre"] = mc

#     # Clean up the "particulars" column
#     df["particulars"] = df["particulars"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)
#     df["voucher_no"] = df["voucher_no"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)

#     df["currency"] = df['excel_row'].map(formats_dict)
    
#     df["currency"] = df["currency"].apply(get_currency_code)

#     if material_centre_name == 10000:
#         df["currency"]  = "USD"
#     if material_centre_name == 92021:
#         df["currency"] = 'GBP'

#     from database.db_crud import DatabaseCrud
#     db_crud = DatabaseCrud(kbe_connector)
    
#     df["exchange_rate"] = df["currency"].apply(lambda x: db_crud.get_exchange_rate_from_db(x)).fillna(0)

# # Calculate amount in INR by multiplying amount by exchange_rate
#     df["amount_in_INR"] = (df["amount"] * df["exchange_rate"]).round(2)    

#     # Drop the excel_row column as it's no longer needed
#     df = df.drop('excel_row', axis=1)

#     return df



# def apply_receivables_transformation(file_path, material_centre_name) -> pd.DataFrame:
    try:
        df = pd.read_excel(file_path, header=None, skipfooter=1)
        date_index = df.loc[df.eq('Date').any(axis=1)].index[0]
        df = df.drop(range(date_index)).reset_index(drop=True)
        df.columns = df.iloc[0]
        df = df.drop(index=0)
        df = df.iloc[1:].reset_index(drop=True)
        
    except FileNotFoundError as e:
        logger.warning(f"Excel File not found in the given {file_path}: {e}")
    if df.empty:
        logger.warning(f"Empty Excel File of {get_compname_tally(file_path)} and report {get_filename_tally(file_path)}")
        return None
    df.columns = df.columns.str.lower().str.replace(" ", "_").str.replace(".", "")
    mc = receivables_comp_codes[int(material_centre_name)]

    df = df.rename(columns={'date': 'voucher_date', 'ref_no': 'voucher_no', 
                            "party's_name": "particulars", 'opening': 'opening_amt', 
                            'pending': 'pending_amt', 'due_on': 'due_date', 
                            'overdue': 'overdue_days', 
                            })
    df["date"] = get_date_tally(path=file_path)
    df['date'] = df['date'].str.removesuffix('.xlsx')
    date_columns = ['date', 'voucher_date', 'due_date']
    for col in date_columns:
         df[col] = pd.to_datetime(df[col], dayfirst=True)

    df["material_centre"] = material_centre_name

    df["particulars"] = df["particulars"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)
    df["voucher_no"] = df["voucher_no"].str.replace('\n', '', regex=True).str.replace('_x000D_', '', regex=True)

    return df


class TallyDataProcessor:
    def __init__(self, excel_file_path) -> None:
        self.excel_file_path = excel_file_path      

    
    def clean_and_transform(self):
        df = None

        company_code = get_compname_tally(self.excel_file_path)
        report_type = get_filename_tally(self.excel_file_path)
        
        if report_type in ['sales', 'sales-return', 'purchase', 'purchase-return']:
            df = apply_transformation(file_path=self.excel_file_path, material_centre_name=company_code)
        
        elif report_type in ['receipts', 'payments', 'journal']:
            df = apply_register_transformation(file_path=self.excel_file_path, material_centre_name=company_code)
        
        elif report_type == "accounts":
            df = apply_accounts_transformation(file_path=self.excel_file_path, material_centre_name=company_code)
        
        elif report_type in ['sales-detailed', 'sales-return-detailed']:
             df = apply_transformation_detailed(file_path=self.excel_file_path, material_centre_name=company_code)

    

        # elif report_type == "items":
        #     df = apply_items_transformation(file_path=self.excel_file_path)

        # elif report_type == "outstanding":
        #     df = apply_outstanding_balance_transformation(file_path=self.excel_file_path, material_centre_name=company_code)

        # # elif report_type == "kbe_outstanding":
        # #     df = apply_kbe_outstanding_transformation(file_path=self.excel_file_path, material_centre_name=company_code)
        
        # elif report_type == "kbe_accounts":
        #     df = apply_kbe_accounts_transformation(file_path=self.excel_file_path, material_centre_name=company_code)

        # elif report_type == "receivables":
        #     df = apply_receivables_transformation(file_path=self.excel_file_path, material_centre_name=company_code)
        
        if df is None:
            logger.error("Dataframe is None!")
            return None

        return df



